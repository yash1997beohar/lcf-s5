#!/usr/bin/env python3
"""
Laxmi Chit Fund - Season 5 : FPL prize tracker.

Pulls the mini-league from the official FPL API, computes every prize standing,
and writes:
    docs/index.html   -> the dashboard (published by GitHub Pages)
    docs/data.json    -> the underlying data (also handy for debugging)
    output/Laxmi_Chit_Fund_S5_audit.xlsx -> the cash-league audit workbook
    snapshots/ownership_gwNN.json -> ownership+status captured at each deadline

Runtime: GitHub Actions (open internet). Local sandboxes may be firewalled from
the FPL API, so use --demo to preview the dashboard with simulated data.
"""

import argparse, json, os, sys, time, datetime as dt, random
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError

API = "https://fantasy.premierleague.com/api"
LEAGUE_ID = 1035071
DIFF_OWNERSHIP_MAX = 7.5          # Differential Diamond threshold (%)
PITY_MAX_GW = 30                  # Pity counts only up to this GW
PITY_MAX_HIT = 8                  # max -hit allowed for a valid Pity score
PITY_MAX_RULED_OUT = 2            # max ruled-out starters allowed
COMEBACK_LOCK_GW = 19            # eligibility frozen after this GW
COMEBACK_BOTTOM_N = 21
TOP5_EXCLUDED = 5                 # Captaincy King / Green Arrow King exclude top 5

NAVY, GOLD, GREY, LIGHT = "#1F3864", "#B7912B", "#595959", "#EDF0F7"

# ---------------------------------------------------------------- HTTP layer
def _get(url, tries=4):
    last = None
    for i in range(tries):
        try:
            req = Request(url, headers={"User-Agent": "Mozilla/5.0 (LCF-tracker)"})
            with urlopen(req, timeout=30) as r:
                return json.load(r)
        except (URLError, HTTPError) as e:
            last = e
            time.sleep(1.5 * (i + 1))
    raise RuntimeError(f"GET failed after {tries} tries: {url} ({last})")

def _cached(cache_dir, key, url):
    """Cache immutable finished-GW responses to disk to keep API calls low."""
    if cache_dir:
        os.makedirs(cache_dir, exist_ok=True)
        path = os.path.join(cache_dir, key + ".json")
        if os.path.exists(path):
            return json.load(open(path))
    d = _get(url)
    if cache_dir:
        json.dump(d, open(path, "w"))
    return d

def fetch_all(league_id, cache_dir="cache"):
    """Pull everything needed. Returns a raw dict. Finished-GW data is cached."""
    boot = _get(f"{API}/bootstrap-static/")
    events = boot["events"]
    # A GW only counts once FPL confirms it: data_checked flips true AFTER bonus points
    # are finalised (finished can flip earlier, before bonus). This guarantees no
    # provisional/pre-bonus scores are ever shown or cached as immutable.
    finished = [e["id"] for e in events if e.get("data_checked")]
    current = next((e["id"] for e in events if e["is_current"]), finished[-1] if finished else 0)
    bonus_pending = any(e.get("finished") and not e.get("data_checked") for e in events)

    # roster: after GW1 members live in standings.results; pre-season in new_entries
    members, page = {}, 1
    while True:
        d = _get(f"{API}/leagues-classic/{league_id}/standings/?page_standings={page}")
        for row in d["standings"]["results"]:
            members[row["entry"]] = {"entry": row["entry"], "team_name": row["entry_name"],
                                     "manager": row["player_name"], "rank": row["rank"],
                                     "last_rank": row["last_rank"], "total": row["total"],
                                     "event_total": row["event_total"]}
        if page == 1:
            league_name = d["league"]["name"]
            for row in d.get("new_entries", {}).get("results", []):
                members.setdefault(row["entry"], {
                    "entry": row["entry"], "team_name": row["entry_name"],
                    "manager": f'{row["player_first_name"]} {row["player_last_name"]}',
                    "rank": 0, "last_rank": 0, "total": 0, "event_total": 0})
        if not d["standings"]["has_next"]:
            break
        page += 1

    histories, picks = {}, {}
    for eid in members:
        try:
            histories[eid] = _get(f"{API}/entry/{eid}/history/")
        except RuntimeError:
            histories[eid] = {"current": [], "chips": []}
        picks[eid] = {}
        for gw in finished:
            try:
                picks[eid][gw] = _cached(cache_dir, f"picks_{eid}_{gw}", f"{API}/entry/{eid}/event/{gw}/picks/")
            except RuntimeError:
                pass

    live = {}
    for gw in finished:
        try:
            d = _cached(cache_dir, f"live_{gw}", f"{API}/event/{gw}/live/")
            live[gw] = {e["id"]: {"p": e["stats"]["total_points"], "m": e["stats"]["minutes"]}
                        for e in d["elements"]}
        except RuntimeError:
            live[gw] = {}

    return {"league_name": league_name, "events": events, "finished": finished,
            "current": current, "elements": boot["elements"], "members": members,
            "histories": histories, "picks": picks, "live": live, "bonus_pending": bonus_pending}

# ------------------------------------------------------------- snapshot (deadline)
def capture_snapshot(elements, gw, snap_dir):
    """Save ownership + availability for a GW (call at/after its deadline)."""
    os.makedirs(snap_dir, exist_ok=True)
    path = os.path.join(snap_dir, f"ownership_gw{gw:02d}.json")
    if os.path.exists(path):
        return
    data = {str(e["id"]): {"own": float(e["selected_by_percent"]),
                            "status": e["status"],
                            "chance": e.get("chance_of_playing_this_round")}
            for e in elements}
    json.dump(data, open(path, "w"))

def load_snapshot(gw, snap_dir):
    path = os.path.join(snap_dir, f"ownership_gw{gw:02d}.json")
    return json.load(open(path)) if os.path.exists(path) else None

# ------------------------------------------------------------------- compute
def _elname(elements):
    return {e["id"]: e["web_name"] for e in elements}

def compute(raw, snap_dir, mt_config):
    M = raw["members"]; finished = raw["finished"]; hist = raw["histories"]
    picks = raw["picks"]; live = raw["live"]; names = _elname(raw["elements"])
    ordered = list(M.values())

    # net GW score per manager per gw
    net = {eid: {} for eid in M}
    for eid in M:
        for row in hist.get(eid, {}).get("current", []):
            net[eid][row["event"]] = row["points"] - row["event_transfers_cost"]
    # transfer hits taken (points lost) per manager
    hits_total = {eid: sum(r.get("event_transfers_cost", 0) for r in hist.get(eid, {}).get("current", []))
                  for eid in M}

    # live player helpers (points / minutes) with captain->vice fallback
    def _pv(gw, el):
        v = live.get(gw, {}).get(el)
        return v if isinstance(v, dict) else ({"p": v or 0, "m": 0} if v is not None else {"p": 0, "m": 0})
    def P(gw, el):   return _pv(gw, el)["p"]
    def MIN(gw, el): return _pv(gw, el)["m"]
    def eff_captain(eid, gw):
        """Captain points (x multiplier). If the captain played 0 mins, the armband
        (and its multiplier, incl. Triple Captain x3) passes to the vice-captain."""
        p = picks.get(eid, {}).get(gw)
        if not p:
            return 0
        cap = next((pk for pk in p["picks"] if pk.get("is_captain")), None)
        vice = next((pk for pk in p["picks"] if pk.get("is_vice_captain")), None)
        if not cap:
            return 0
        mult = cap["multiplier"]
        if MIN(gw, cap["element"]) > 0:
            return P(gw, cap["element"]) * mult
        if vice and MIN(gw, vice["element"]) > 0:
            return P(gw, vice["element"]) * mult
        return 0

    # -------- Overall table
    if any(m["rank"] for m in M.values()):
        overall = sorted(ordered, key=lambda m: (m["rank"] if m["rank"] else 1e9))
    else:
        overall = sorted(ordered, key=lambda m: -m["total"])
    overall_tbl = []
    for i, m in enumerate(overall, 1):
        lr, r = m["last_rank"], (m["rank"] or i)
        arrow = "•" if not lr or lr == r else ("▲" if r < lr else "▼")
        overall_tbl.append({"rank": r or i, "arrow": arrow, "manager": m["manager"],
                            "team_name": m["team_name"], "entry": m["entry"],
                            "gw_points": m["event_total"], "total": m["total"],
                            "hits": hits_total.get(m["entry"], 0)})
    top5_ids = {row["entry"] for row in overall_tbl[:TOP5_EXCLUDED]}

    # -------- GW score matrix
    matrix = {"gws": finished,
              "rows": [{"manager": M[eid]["manager"], "entry": eid,
                        "scores": [net[eid].get(gw) for gw in finished]} for eid in M]}

    # -------- Chip Kings
    chip_key = {"3xc": "Triple Captain", "bboost": "Bench Boost",
                "wildcard": "Wildcard", "freehit": "Free Hit"}
    chip_best = {k: {"score": None, "manager": None, "gw": None} for k in chip_key}
    for eid in M:
        for c in hist.get(eid, {}).get("chips", []):
            name, gw = c["name"], c["event"]
            if name not in chip_key or gw not in finished:
                continue
            if name == "3xc":
                score = eff_captain(eid, gw)   # x3 already in the captain's multiplier; vice fallback applied
            else:
                score = net[eid].get(gw)
            if score is None:
                continue
            b = chip_best[name]
            if b["score"] is None or score > b["score"]:
                chip_best[name] = {"score": score, "manager": M[eid]["manager"], "gw": gw}
    chip_kings = [{"chip": chip_key[k], "code": k, **chip_best[k]} for k in chip_key]

    # -------- per-GW captain & bench points (used by Captaincy King + MT knockout tiebreaks)
    cap_gw = {eid: {} for eid in M}
    bench_gw = {eid: {} for eid in M}
    for eid in M:
        for r in hist.get(eid, {}).get("current", []):
            bench_gw[eid][r["event"]] = r.get("points_on_bench", 0)
        for gw in finished:
            if picks.get(eid, {}).get(gw):
                cap_gw[eid][gw] = eff_captain(eid, gw)
    cap_tot = {eid: sum(cap_gw[eid].values()) for eid in M}
    captaincy = sorted(
        [{"manager": M[eid]["manager"], "entry": eid, "points": cap_tot[eid],
          "excluded": eid in top5_ids} for eid in M],
        key=lambda x: -x["points"])

    arrows = {eid: 0 for eid in M}
    for eid in M:
        cur = {r["event"]: r["overall_rank"] for r in hist.get(eid, {}).get("current", [])}
        for gw in finished:
            if gw >= 2 and cur.get(gw) and cur.get(gw - 1) and cur[gw] < cur[gw - 1]:
                arrows[eid] += 1
    green = sorted(
        [{"manager": M[eid]["manager"], "entry": eid, "arrows": arrows[eid],
          "excluded": eid in top5_ids} for eid in M],
        key=lambda x: -x["arrows"])

    # -------- Comeback (locked after GW19)
    comeback = {"active": (raw["current"] or 0) > COMEBACK_LOCK_GW, "board": [], "eligible_n": COMEBACK_BOTTOM_N}
    if comeback["active"]:
        # rank after GW19 from overall_rank in league? use net-total to GW19
        tot19 = {eid: sum(v for g, v in net[eid].items() if g <= COMEBACK_LOCK_GW) for eid in M}
        eligible = sorted(M, key=lambda e: tot19[e])[:COMEBACK_BOTTOM_N]
        board = [{"manager": M[eid]["manager"], "entry": eid,
                  "points": sum(v for g, v in net[eid].items() if g > COMEBACK_LOCK_GW)}
                 for eid in eligible]
        comeback["board"] = sorted(board, key=lambda x: -x["points"])

    # -------- Differential Diamond
    diff_board = []
    for gw in finished:
        snap = load_snapshot(gw, snap_dir)
        if not snap:
            continue
        for eid in M:
            p = picks.get(eid, {}).get(gw)
            if not p:
                continue
            for pk in p["picks"]:
                if pk["multiplier"] <= 0:      # starting XI only
                    continue
                s = snap.get(str(pk["element"]))
                if s and s["own"] < DIFF_OWNERSHIP_MAX:
                    diff_board.append({"manager": M[eid]["manager"], "gw": gw,
                                       "player": names.get(pk["element"], "?"),
                                       "points": P(gw, pk["element"]),
                                       "own": s["own"]})
    diff_board.sort(key=lambda x: -x["points"])
    differential = {"best": diff_board[0] if diff_board else None, "board": diff_board[:15],
                    "has_snapshots": any(load_snapshot(g, snap_dir) for g in finished)}

    # -------- Pity the Living Dead (candidates; admin confirms validity)
    pity = []
    for eid in M:
        for gw in finished:
            if gw > PITY_MAX_GW or gw not in net[eid]:
                continue
            row = next((r for r in hist[eid]["current"] if r["event"] == gw), {})
            hit = row.get("event_transfers_cost", 0)
            ruled = None
            snap = load_snapshot(gw, snap_dir)
            p = picks.get(eid, {}).get(gw)
            if snap and p:
                ruled = sum(1 for pk in p["picks"] if pk["multiplier"] > 0
                            and snap.get(str(pk["element"]), {}).get("status") in ("i", "s", "u", "n"))
            valid = (hit <= PITY_MAX_HIT) and (ruled is None or ruled <= PITY_MAX_RULED_OUT)
            pity.append({"manager": M[eid]["manager"], "gw": gw, "score": net[eid][gw],
                         "hit": hit, "ruled_out": ruled, "valid": valid})
    pity.sort(key=lambda x: x["score"])
    pity = pity[:12]

    # -------- Manager Profiles (from FPL past-season history)
    league_rank = {row["entry"]: row["rank"] for row in overall_tbl}
    profiles = []
    for eid, m in M.items():
        past = hist.get(eid, {}).get("past", []) or []
        best = min((p["rank"] for p in past), default=None)
        bseason = min(past, key=lambda p: p["rank"])["season_name"] if past else None
        profiles.append({"manager": m["manager"], "team_name": m["team_name"],
                         "league_rank": league_rank.get(eid, 0), "seasons": len(past) + 1,
                         "best_rank": best, "best_season": bseason,
                         "last_rank": past[-1]["rank"] if past else None})
    profiles.sort(key=lambda x: x["league_rank"] or 1e9)

    # -------- Season Stats (per manager)
    gw_avg = {}
    for g in finished:
        xs = [net[eid][g] for eid in M if g in net[eid]]
        gw_avg[g] = (sum(xs) / len(xs)) if xs else 0
    stats = []
    for eid, m in M.items():
        cur = hist.get(eid, {}).get("current", [])
        scores = [(net[eid][g], g) for g in finished if g in net[eid]]
        vals = [s for s, _ in scores]
        hi = max(scores, default=(None, None))
        lo = min(scores, default=(None, None))
        above_avg = sum(1 for g in finished if g in net[eid] and net[eid][g] > gw_avg[g])
        orank = {r["event"]: r.get("overall_rank") for r in cur}
        best_jump, best_jump_gw = None, None
        for g in finished:
            if g >= 2 and orank.get(g) and orank.get(g - 1):
                jump = orank[g - 1] - orank[g]          # positive = climbed
                if jump > 0 and (best_jump is None or jump > best_jump):
                    best_jump, best_jump_gw = jump, g
        v0 = cur[0].get("value") if cur else None
        v1 = cur[-1].get("value") if cur else None
        stats.append({"manager": m["manager"], "team_name": m["team_name"], "entry": eid,
                      "total": m["total"], "played": len(vals),
                      "avg": round(sum(vals) / len(vals), 1) if vals else None,
                      "above_avg": above_avg, "cap_return": cap_tot.get(eid, 0),
                      "high": hi[0], "high_gw": hi[1], "low": lo[0], "low_gw": lo[1],
                      "transfers": sum(r.get("event_transfers", 0) for r in cur),
                      "hits_pts": hits_total[eid], "hits_n": hits_total[eid] // 4,
                      "bench": sum(r.get("points_on_bench", 0) for r in cur),
                      "value": round(v1 / 10.0, 1) if v1 is not None else None,
                      "value_delta": round((v1 - v0) / 10.0, 1) if (v0 is not None and v1 is not None) else None,
                      "bank": round(cur[-1].get("bank", 0) / 10.0, 1) if cur else None,
                      "grank": cur[-1].get("overall_rank") if cur else None,
                      "best_jump": best_jump, "best_jump_gw": best_jump_gw,
                      "chips": [c["name"] for c in hist.get(eid, {}).get("chips", [])]})
    stats.sort(key=lambda x: -x["total"])

    # -------- per-GW series (for rank chart, GW-score bars, H2H) — pure data export
    series = {}
    for eid in M:
        cur = {r["event"]: r for r in hist.get(eid, {}).get("current", [])}
        series[str(eid)] = [{"gw": g, "net": net[eid][g], "orank": cur.get(g, {}).get("overall_rank"),
                             "cap": cap_gw[eid].get(g, 0)} for g in finished if g in net[eid]]

    # -------- Mini Tournaments (draws come from config)
    mini = compute_mts(mt_config, M, net, finished, cap_gw, bench_gw)

    return {"meta": {"league_name": raw["league_name"], "league_id": LEAGUE_ID,
                     "generated_utc": dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
                     "generated_iso": dt.datetime.utcnow().isoformat() + "Z",
                     "started": (raw["current"] or 0) > 0 or bool(finished),
                     "current_gw": raw["current"], "finished": finished,
                     "n_members": len(M), "xlsx": "Laxmi_Chit_Fund_S5_audit.xlsx",
                     "bonus_pending": raw.get("bonus_pending", False)},
            "overall": overall_tbl, "matrix": matrix, "chip_kings": chip_kings,
            "captaincy": captaincy, "green": green, "comeback": comeback,
            "differential": differential, "pity": pity, "profiles": profiles,
            "stats": stats, "series": series, "mini": mini}

def compute_mts(mt_config, M, net, finished, cap_gw, bench_gw):
    """mt_config: list of {name, group_gws, ko_gws, groups:{A:[eids]...}}. Optional."""
    out = []
    id2name = {eid: M[eid]["manager"] for eid in M}
    for mt in mt_config or []:
        groups = mt.get("groups") or {}
        gtables, standings = {}, {}
        for gname, eids in groups.items():
            recs = {e: {"P": 0, "W": 0, "D": 0, "L": 0, "pts": 0, "fpl": 0} for e in eids}
            sched = _rr4(eids)
            for idx, gw in enumerate(mt["group_gws"]):
                if gw not in finished or idx >= len(sched):
                    continue
                for a, b in sched[idx]:
                    sa, sb = net.get(a, {}).get(gw), net.get(b, {}).get(gw)
                    if sa is None or sb is None:
                        continue
                    recs[a]["P"] += 1; recs[b]["P"] += 1
                    recs[a]["fpl"] += sa; recs[b]["fpl"] += sb
                    if sa > sb: recs[a]["W"] += 1; recs[a]["pts"] += 3; recs[b]["L"] += 1
                    elif sb > sa: recs[b]["W"] += 1; recs[b]["pts"] += 3; recs[a]["L"] += 1
                    else: recs[a]["D"] += 1; recs[b]["D"] += 1; recs[a]["pts"] += 1; recs[b]["pts"] += 1
            order = sorted(eids, key=lambda e: (-recs[e]["pts"], -recs[e]["fpl"]))
            standings[gname] = [(e, recs[e]) for e in order]
            gtables[gname] = [[id2name[e], recs[e]] for e in order]

        knockout = None
        if groups and all(g in finished for g in mt["group_gws"]):
            knockout = _build_knockout(standings, mt["ko_gws"], finished, net, cap_gw, bench_gw, id2name)

        out.append({"name": mt["name"], "group_gws": mt["group_gws"], "ko_gws": mt["ko_gws"],
                    "groups": gtables, "has_draw": bool(groups),
                    "played": [g for g in mt["group_gws"] if g in finished], "knockout": knockout})
    return out

def _ko_winner(a, b, gw, finished, net, cap_gw, bench_gw, seed):
    """Rulebook tiebreak: net score -> captain pts -> bench pts -> better seed."""
    sa, sb = net.get(a, {}).get(gw), net.get(b, {}).get(gw)
    if a is None or b is None or gw not in finished or sa is None or sb is None:
        return None, sa, sb
    if sa != sb:
        return ("a" if sa > sb else "b"), sa, sb
    for tb in (cap_gw, bench_gw):
        va, vb = tb.get(a, {}).get(gw, 0), tb.get(b, {}).get(gw, 0)
        if va != vb:
            return ("a" if va > vb else "b"), sa, sb
    return ("a" if seed.get(a, 99) < seed.get(b, 99) else "b"), sa, sb

def _build_knockout(standings, ko_gws, finished, net, cap_gw, bench_gw, id2name):
    winners = sorted([standings[g][0] for g in standings], key=lambda x: (-x[1]["pts"], -x[1]["fpl"]))
    runners = sorted([standings[g][1] for g in standings if len(standings[g]) > 1],
                     key=lambda x: (-x[1]["pts"], -x[1]["fpl"]))
    seeds = [w[0] for w in winners[:6]] + [r[0] for r in runners[:2]]
    if len(seeds) < 8:
        return None
    seed = {e: i + 1 for i, e in enumerate(seeds)}
    s = seeds
    gw_qf, gw_sf, gw_f = (list(ko_gws) + [None, None, None])[:3]
    nm = lambda e: id2name.get(e, "—") if e else "—"

    def round_of(pairs, gw, label):
        res, ties = [], []
        for a, b in pairs:
            w, sa, sb = _ko_winner(a, b, gw, finished, net, cap_gw, bench_gw, seed)
            res.append(a if w == "a" else (b if w == "b" else None))
            ties.append({"a": nm(a), "b": nm(b), "sa": sa, "sb": sb, "winner": w})
        return res, {"name": label, "ties": ties}

    qf_res, qf = round_of([(s[0], s[7]), (s[3], s[4]), (s[2], s[5]), (s[1], s[6])], gw_qf, "Quarterfinals")
    sf_res, sf = round_of([(qf_res[0], qf_res[1]), (qf_res[2], qf_res[3])], gw_sf, "Semifinals")
    f_res, fin = round_of([(sf_res[0], sf_res[1])], gw_f, "Final")
    return {"rounds": [qf, sf, fin],
            "seeds": [{"seed": i + 1, "manager": nm(e)} for i, e in enumerate(seeds)],
            "champion": nm(f_res[0]) if f_res[0] else None}

def _rr4(t):
    """3-round round-robin schedule for a group of 4 (indices into t)."""
    if len(t) != 4:
        return []
    a, b, c, d = t
    return [[(a, b), (c, d)], [(a, c), (b, d)], [(a, d), (b, c)]]

# --------------------------------------------------------------------- render
def render_html(data):
    j = json.dumps(data).replace("</", "<\\/")
    here = os.path.dirname(os.path.abspath(__file__))
    tpath = os.path.join(here, "template.html")
    if os.path.exists(tpath):                       # external template (preferred)
        return open(tpath, encoding="utf-8").read().replace("__DATA__", j)
    return HTML_TEMPLATE.replace("__NAVY__", NAVY).replace("__GOLD__", GOLD)\
        .replace("__GREY__", GREY).replace("__LIGHT__", LIGHT).replace("__DATA__", j)  # legacy fallback

# --------------------------------------------------------------------- excel
def write_excel(data, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl not installed; skipping Excel", file=sys.stderr); return
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb = Workbook()
    hdr = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="1F3864")
    def style(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(1, c); cell.font = hdr; cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
    ws = wb.active; ws.title = "Overall"
    ws.append(["Rank", "Manager", "Team", "GW", "Hits (pts)", "Total"])
    for r in data["overall"]:
        ws.append([r["rank"], r["manager"], r["team_name"], r["gw_points"], -r.get("hits", 0), r["total"]])
    style(ws, 6)
    # GW matrix
    ws = wb.create_sheet("GW Scores")
    ws.append(["Manager"] + [f"GW{g}" for g in data["matrix"]["gws"]])
    for row in data["matrix"]["rows"]:
        ws.append([row["manager"]] + [s if s is not None else "" for s in row["scores"]])
    style(ws, 1 + len(data["matrix"]["gws"]))
    # side prizes
    ws = wb.create_sheet("Chip Kings")
    ws.append(["Chip", "Leader", "Score", "GW"])
    for c in data["chip_kings"]:
        ws.append([c["chip"], c["manager"] or "-", c["score"] if c["score"] is not None else "-", c["gw"] or "-"])
    style(ws, 4)
    ws = wb.create_sheet("Captaincy King")
    ws.append(["Manager", "Captain Pts", "Excluded (top 5)"])
    for r in data["captaincy"]:
        ws.append([r["manager"], r["points"], "yes" if r["excluded"] else ""])
    style(ws, 3)
    ws = wb.create_sheet("Green Arrow King")
    ws.append(["Manager", "Green Arrows", "Excluded (top 5)"])
    for r in data["green"]:
        ws.append([r["manager"], r["arrows"], "yes" if r["excluded"] else ""])
    style(ws, 3)
    ws = wb.create_sheet("Comeback")
    ws.append(["Manager", "Pts since GW20"])
    for r in data["comeback"]["board"]:
        ws.append([r["manager"], r["points"]])
    style(ws, 2)
    ws = wb.create_sheet("Differential Diamond")
    ws.append(["Manager", "Player", "GW", "Points", "Own %"])
    for r in data["differential"]["board"]:
        ws.append([r["manager"], r["player"], r["gw"], r["points"], r["own"]])
    style(ws, 5)
    ws = wb.create_sheet("Profiles")
    ws.append(["League #", "Manager", "Team", "Seasons", "Best Rank", "Best Season", "Last Season Rank"])
    for r in data.get("profiles", []):
        ws.append([r["league_rank"], r["manager"], r["team_name"], r["seasons"],
                   r["best_rank"], r["best_season"], r["last_rank"]])
    style(ws, 7)
    ws = wb.create_sheet("Season Stats")
    ws.append(["Manager", "Team", "Total", "Avg GW", "Above Avg", "Cap Return", "High", "High GW",
               "Low", "Low GW", "Transfers", "Hits (n)", "Hits (pts)", "Bench", "Squad £m",
               "Value Δ £m", "Bank £m", "Global Rank", "Best Jump", "Best Jump GW", "Chips"])
    for r in data.get("stats", []):
        ws.append([r["manager"], r["team_name"], r["total"], r["avg"], r.get("above_avg"),
                   r.get("cap_return"), r["high"], r["high_gw"], r["low"], r["low_gw"], r["transfers"],
                   r["hits_n"], -r["hits_pts"], r["bench"], r["value"], r.get("value_delta"),
                   r["bank"], r["grank"], r.get("best_jump"), r.get("best_jump_gw"), ", ".join(r["chips"])])
    style(ws, 21)
    ws = wb.create_sheet("Pity")
    ws.append(["Manager", "GW", "Score", "Hit", "Ruled-out starters", "Valid?"])
    for r in data["pity"]:
        ws.append([r["manager"], r["gw"], r["score"], -r["hit"], r["ruled_out"] if r["ruled_out"] is not None else "n/a",
                   "yes" if r["valid"] else "review"])
    style(ws, 6)
    wb.save(path)

# ------------------------------------------------------------------ demo data
def demo_data():
    roster = [(1301167,"Cment_mixers","Deepankar Sachdeva"),(5419891,"Onside Offside","Ajith Thomas"),
        (6023506,"sakalaka","prateek gupta"),(5970777,"fpl_26_27","Karan Chopra"),
        (5926399,"BattleBorn","Aquib Razack"),(46711,"Xa bing chilling","Satweek Nayak"),
        (1086057,"ShakeNBake","Arshaq Razack"),(5582398,"Southall United","Rohan Agrawal"),
        (5565611,"Slumber FC","Dean Jenkins"),(5008941,"Pole Calmer","Ashwin Jose"),
        (2424090,"Cold Palmer","Shivanshu Madan"),(5340498,"Old Reijnders Nagar","Sid Shaurya"),
        (3330149,"Amad Aura","Karan Khanna"),(79306,"Totally Szobo","Abhinav Premsekhar"),
        (3323226,"ItDzntMata","Danish Javed"),(4993853,"Redsbay","Nilay jain"),
        (256447,"Fiery Raiders","Anish Joshi"),(3942416,"115 FC","Abhay Rawat"),
        (2354537,"MaSalah","Raveesh Kalra"),(427154,"KingOfKings Fc","Abhiraj Singh"),
        (3909562,"Goal Diggers FC","Varun Khurana"),(4837220,"FC Teetotallers","Rajit Das"),
        (3181909,"Mishra's Thirteen","Aparnesh Mishra"),(2003955,"Loading...","Aayush Khanna"),
        (4688717,"Enzotic","Abhishek Gupte"),(20569,"Captn Magnifico","Naman Agrawal"),
        (1362641,"Carrick Janta Party","Pranshul Jain"),(9695,"His Name is Diogo","Aditya Vaidya"),
        (374875,"Fluente in Football","Yash Beohar")]
    rng = random.Random(7); GWS = list(range(1, 8))
    players = ["Haaland","Salah","Palmer","Saka","Isak","Watkins","Mbeumo","Gordon","Cunha","Rogers",
               "Mateta","Wood","Semenyo","Muniz","Sarr","Rice","Gakpo","Foden","Bruno","Solanke"]
    net = {e[0]: {g: rng.randint(28, 92) for g in GWS} for e in roster}
    members, matrix_rows = {}, []
    for e in roster:
        eid, tn, mn = e
        tot = sum(net[eid].values())
        members[eid] = {"entry": eid, "team_name": tn, "manager": mn, "total": tot,
                        "event_total": net[eid][GWS[-1]]}
    order = sorted(members, key=lambda x: -members[x]["total"])
    overall = []
    for i, eid in enumerate(order, 1):
        lr = i + rng.choice([-2, -1, 0, 1, 2])
        arrow = "•" if lr == i else ("▲" if i < lr else "▼")
        overall.append({"rank": i, "arrow": arrow, "manager": members[eid]["manager"],
                        "team_name": members[eid]["team_name"], "entry": eid,
                        "gw_points": members[eid]["event_total"], "total": members[eid]["total"],
                        "hits": rng.choice([0, 0, 0, 4, 8, 12])})
        matrix_rows.append({"manager": members[eid]["manager"], "entry": eid,
                            "scores": [net[eid][g] for g in GWS]})
    top5 = {o["entry"] for o in overall[:5]}
    chips = [("Triple Captain","3xc"),("Bench Boost","bboost"),("Wildcard","wildcard"),("Free Hit","freehit")]
    chip_kings = [{"chip": c, "code": k, "manager": members[rng.choice(order)]["manager"],
                   "score": rng.randint(24, 39) if k=="3xc" else rng.randint(78, 121),
                   "gw": rng.choice(GWS)} for c, k in chips]
    captaincy = sorted([{"manager": members[e]["manager"], "entry": e,
                         "points": rng.randint(48, 132), "excluded": e in top5} for e in order],
                       key=lambda x: -x["points"])
    green = sorted([{"manager": members[e]["manager"], "entry": e,
                     "arrows": rng.randint(0, 6), "excluded": e in top5} for e in order],
                   key=lambda x: -x["arrows"])
    diff = sorted([{"manager": members[rng.choice(order)]["manager"], "gw": rng.choice(GWS),
                    "player": rng.choice(players), "points": rng.randint(9, 21),
                    "own": round(rng.uniform(1.2, 7.4), 1)} for _ in range(14)],
                  key=lambda x: -x["points"])
    pity = sorted([{"manager": members[rng.choice(order)]["manager"], "gw": rng.choice(GWS),
                    "score": rng.randint(12, 34), "hit": rng.choice([0,0,0,4,8]),
                    "ruled_out": rng.choice([0,0,1,2]), "valid": True} for _ in range(12)],
                  key=lambda x: x["score"])[:12]
    gtables = _demo_groups(order, members, net, rng)
    mini = [{"name": "MT1 (GW3–GW8)", "group_gws": [3,4,5], "ko_gws": [6,7,8], "has_draw": True,
             "played": [3,4,5], "groups": gtables, "knockout": _demo_knockout(gtables, rng)}]
    seasons_pool = ["2019/20","2020/21","2021/22","2022/23","2023/24","2024/25","2025/26"]
    profiles = [{"manager": o["manager"], "team_name": o["team_name"], "league_rank": o["rank"],
                 "seasons": rng.randint(2, 8), "best_rank": rng.randint(4000, 480000),
                 "best_season": rng.choice(seasons_pool), "last_rank": rng.randint(20000, 1200000)}
                for o in overall]
    chips_pool = [["3xc"], ["bboost"], ["wildcard"], [], ["freehit"], ["wildcard", "3xc"]]
    stats = []
    for o in overall:
        eid = o["entry"]; sc = [(net[eid][g], g) for g in GWS]
        vals = [v for v, _ in sc]; hi = max(sc); lo = min(sc)
        stats.append({"manager": o["manager"], "team_name": o["team_name"], "entry": eid,
                      "total": o["total"], "played": len(GWS), "avg": round(sum(vals) / len(vals), 1),
                      "above_avg": rng.randint(0, len(GWS)), "cap_return": rng.randint(40, 130),
                      "high": hi[0], "high_gw": hi[1], "low": lo[0], "low_gw": lo[1],
                      "transfers": rng.randint(7, 22), "hits_pts": o["hits"], "hits_n": o["hits"] // 4,
                      "bench": rng.randint(20, 80), "value": round(rng.uniform(99.5, 103.5), 1),
                      "value_delta": round(rng.uniform(-1.5, 3.0), 1),
                      "bank": round(rng.uniform(0, 2.5), 1), "grank": rng.randint(50000, 3000000),
                      "best_jump": rng.randint(50000, 2500000), "best_jump_gw": rng.choice(GWS),
                      "chips": rng.choice(chips_pool)})
    series = {}
    for o in overall:
        eid = o["entry"]; orank = rng.randint(300000, 3000000); s = []
        for g in GWS:
            orank = max(800, orank - rng.randint(-90000, 240000))
            s.append({"gw": g, "net": net[eid][g], "orank": orank, "cap": rng.randint(2, 30)})
        series[str(eid)] = s
    return {"meta": {"league_name": "Laxmi Chit Fund - Season 5 (DEMO)", "league_id": LEAGUE_ID,
                     "generated_utc": dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
                     "generated_iso": dt.datetime.utcnow().isoformat() + "Z", "started": True,
                     "current_gw": 7, "finished": GWS, "n_members": len(roster), "demo": True,
                     "xlsx": "S5-Dashboard-audit-DEMO.xlsx"},
            "overall": overall, "series": series, "matrix": {"gws": GWS, "rows": matrix_rows},
            "chip_kings": chip_kings, "captaincy": captaincy, "green": green,
            "comeback": {"active": False, "board": [], "eligible_n": COMEBACK_BOTTOM_N},
            "differential": {"best": diff[0], "board": diff[:15], "has_snapshots": True},
            "pity": pity, "profiles": profiles, "stats": stats, "mini": mini}

def _demo_knockout(gtables, rng):
    q = [tbl[0][0] for tbl in gtables.values()] + [tbl[1][0] for tbl in gtables.values()][:2]
    q = q[:8]; rng.shuffle(q)
    def tie(a, b):
        sa, sb = rng.randint(38, 88), rng.randint(38, 88)
        if sa == sb: sb += 1
        return {"a": a, "b": b, "sa": sa, "sb": sb, "winner": "a" if sa > sb else "b"}
    qf = [tie(q[0], q[1]), tie(q[2], q[3]), tie(q[4], q[5]), tie(q[6], q[7])]
    sfn = [t["a"] if t["winner"] == "a" else t["b"] for t in qf]
    sf = [tie(sfn[0], sfn[1]), tie(sfn[2], sfn[3])]
    ffn = [t["a"] if t["winner"] == "a" else t["b"] for t in sf]
    fin = [tie(ffn[0], ffn[1])]
    champ = fin[0]["a"] if fin[0]["winner"] == "a" else fin[0]["b"]
    return {"rounds": [{"name": "Quarterfinals", "ties": qf}, {"name": "Semifinals", "ties": sf},
                       {"name": "Final", "ties": fin}], "champion": champ}

def _demo_groups(order, members, net, rng):
    q = order[:24]; groups = {}
    for i, g in enumerate("ABCDEF"):
        eids = q[i*4:(i+1)*4]
        recs = []
        for e in eids:
            w = rng.randint(0, 3); d = rng.randint(0, 3 - w)
            recs.append([members[e]["manager"], {"P": 3, "W": w, "D": d, "L": 3 - w - d,
                         "pts": 3*w + d, "fpl": sum(net[e][x] for x in (3,4,5))}])
        groups[g] = sorted(recs, key=lambda r: (-r[1]["pts"], -r[1]["fpl"]))
    return groups

# --------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--demo", action="store_true", help="render with simulated data")
    ap.add_argument("--out", default="docs")
    ap.add_argument("--snap", default="snapshots")
    ap.add_argument("--xlsx", default="docs/Laxmi_Chit_Fund_S5_audit.xlsx")
    ap.add_argument("--config", default="config.json")
    a = ap.parse_args()

    if a.demo:
        data = demo_data()
    else:
        cfg = json.load(open(a.config)) if os.path.exists(a.config) else {}
        raw = fetch_all(cfg.get("league_id", LEAGUE_ID))
        # capture a deadline snapshot for the current GW (ownership won't change post-deadline)
        if raw["current"]:
            capture_snapshot(raw["elements"], raw["current"], a.snap)
        data = compute(raw, a.snap, cfg.get("mini_tournaments", []))

    os.makedirs(a.out, exist_ok=True)
    open(os.path.join(a.out, "index.html"), "w").write(render_html(data))
    json.dump(data, open(os.path.join(a.out, "data.json"), "w"), indent=1)
    write_excel(data, a.xlsx)
    print(f"OK: {data['meta']['n_members']} managers, current GW {data['meta']['current_gw']}")

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Laxmi Chit Fund S5 — Live Standings</title>
<style>
 :root{--navy:__NAVY__;--gold:__GOLD__;--grey:__GREY__;--light:__LIGHT__}
 *{box-sizing:border-box} body{margin:0;font-family:-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;color:#222;background:#f6f7fb}
 header{background:var(--navy);color:#fff;padding:18px 20px}
 header h1{margin:0;font-size:20px;letter-spacing:.5px} header .sub{color:#cdd6ea;font-size:13px;margin-top:3px}
 header .gold{color:var(--gold);font-weight:700}
 .wrap{max-width:1040px;margin:0 auto;padding:0 14px 60px}
 .tabs{display:flex;flex-wrap:wrap;gap:6px;position:sticky;top:0;background:#f6f7fb;padding:12px 0;z-index:5;border-bottom:1px solid #e3e7f0}
 .tab{padding:8px 13px;border:1px solid #d6dcea;border-radius:20px;background:#fff;cursor:pointer;font-size:13px;font-weight:600;color:var(--navy)}
 .tab.active{background:var(--navy);color:#fff;border-color:var(--navy)}
 .panel{display:none;margin-top:16px;animation:f .2s ease} .panel.active{display:block} @keyframes f{from{opacity:0}to{opacity:1}}
 h2{color:var(--navy);font-size:17px;margin:6px 0 4px} .note{color:var(--grey);font-size:12.5px;margin:0 0 12px}
 table{width:100%;border-collapse:collapse;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 1px 3px rgba(20,30,60,.07);font-size:13.5px}
 th{background:var(--navy);color:#fff;text-align:left;padding:9px 11px;font-weight:600} td{padding:8px 11px;border-top:1px solid #eef1f7}
 tr:nth-child(even) td{background:var(--light)} .r{text-align:right} .c{text-align:center}
 .up{color:#2e8b57;font-weight:700} .down{color:#c0392b;font-weight:700} .flat{color:#9aa0ad}
 .badge{display:inline-block;background:var(--gold);color:#fff;border-radius:6px;padding:1px 7px;font-size:11px;font-weight:700}
 .cardrow{display:grid;grid-template-columns:repeat(auto-fit,minmax(210px,1fr));gap:12px;margin-bottom:14px}
 .card{background:#fff;border-radius:10px;padding:13px 15px;box-shadow:0 1px 3px rgba(20,30,60,.07);border-left:5px solid var(--gold)}
 .card .k{font-size:12px;color:var(--grey);text-transform:uppercase;letter-spacing:.4px} .card .v{font-size:22px;font-weight:800;color:var(--navy);margin-top:3px}
 .card .s{font-size:12.5px;color:#333;margin-top:2px}
 .excl{opacity:.5} .top5 td{background:#fbf6e7 !important}
 .grp{display:grid;grid-template-columns:repeat(auto-fit,minmax(300px,1fr));gap:12px}
 .mut{color:var(--grey);font-size:12px} .pill{font-size:11px;background:#e8edf7;color:var(--navy);border-radius:10px;padding:1px 8px;margin-left:6px}
 footer{color:var(--grey);font-size:11.5px;text-align:center;padding:22px}
</style></head><body>
<header><div class="wrap" style="padding:0">
 <h1>🏆 Laxmi Chit Fund — Season 5</h1>
 <div class="sub"><span id="hsub"></span></div></div></header>
<div class="wrap">
 <div class="tabs" id="tabs"></div>
 <div id="panels"></div>
 <footer id="foot"></footer>
</div>
<script>
const D = __DATA__;
const el=(t,c,h)=>{const e=document.createElement(t);if(c)e.className=c;if(h!=null)e.innerHTML=h;return e};
function arrow(a){return a==='▲'?'<span class="up">▲</span>':a==='▼'?'<span class="down">▼</span>':'<span class="flat">•</span>'}
function table(cols,rows){let h='<table><thead><tr>'+cols.map(c=>`<th class="${c.cls||''}">${c.t}</th>`).join('')+'</tr></thead><tbody>';
 rows.forEach(r=>{h+='<tr'+(r._cls?` class="${r._cls}"`:'')+'>'+r.cells.map((c,i)=>`<td class="${cols[i].cls||''}">${c}</td>`).join('')+'</tr>'});return h+'</tbody></table>'}

const TABS=[];
function addTab(id,label,render){TABS.push({id,label,render})}

addTab('overall','Overall',()=>{
 const rows=D.overall.map(r=>({_cls:r.rank<=5?'top5':'',cells:[`${r.rank} ${arrow(r.arrow)}`,r.manager,`<span class="mut">${r.team_name}</span>`,r.gw_points,`<b>${r.total}</b>`]}));
 let h='<h2>Overall Table</h2><p class="note">Top 5 (shaded) share ₹75,000 at GW38. Arrows show movement since last GW.</p>';
 return h+table([{t:'#'},{t:'Manager'},{t:'Team'},{t:'GW',cls:'r'},{t:'Total',cls:'r'}],rows);
});

addTab('mini','Mini Tournaments',()=>{
 if(!D.mini||!D.mini.length) return '<h2>Mini Tournaments</h2><p class="note">The first mini-tournament (MT1) begins at GW3. Standings will appear here once the draw is made after GW2.</p>';
 let h='';
 D.mini.forEach(mt=>{
   h+=`<h2>${mt.name}</h2>`;
   if(!mt.has_draw){h+='<p class="note">Awaiting the random group draw (made after the qualification cut).</p>';return}
   h+=`<p class="note">Group stage GW ${mt.group_gws.join(', ')} · Knockouts GW ${mt.ko_gws.join(', ')} · played: ${mt.played.join(', ')||'—'}</p><div class="grp">`;
   Object.entries(mt.groups).forEach(([g,tbl])=>{
     const rows=tbl.map((r,i)=>({_cls:i<2?'top5':'',cells:[r[0],r[1].P,r[1].W,r[1].D,r[1].L,`<b>${r[1].pts}</b>`,r[1].fpl]}));
     h+='<div>'+`<div class="mut" style="font-weight:700;color:var(--navy);margin:4px 0">Group ${g}</div>`+
        table([{t:'Team'},{t:'P',cls:'c'},{t:'W',cls:'c'},{t:'D',cls:'c'},{t:'L',cls:'c'},{t:'Pts',cls:'c'},{t:'FPL',cls:'r'}],rows)+'</div>';
   });
   h+='</div><p class="note" style="margin-top:8px">Top 2 of each group (shaded) + 2 best runners-up advance to the 8-team knockout.</p>';
 });
 return h;
});

addTab('chip','Chip Kings',()=>{
 let h='<h2>Chip Kings — ₹3,000 each</h2><p class="note">Best single-GW return with each chip (best of both halves).</p><div class="cardrow">';
 D.chip_kings.forEach(c=>{h+=`<div class="card"><div class="k">${c.chip}</div><div class="v">${c.score!=null?c.score:'—'}</div><div class="s">${c.manager?c.manager+' · GW'+c.gw:'not played yet'}</div></div>`});
 return h+'</div>';
});

addTab('captain','Captaincy King',()=>{
 const rows=D.captaincy.map((r,i)=>({_cls:r.excluded?'excl':'',cells:[i+1,r.manager+(r.excluded?' <span class="pill">top-5 excl.</span>':''),`<b>${r.points}</b>`]}));
 return '<h2>Captaincy King — ₹3,000</h2><p class="note">Total captain points (after multipliers). Top-5 overall are excluded at settlement.</p>'+
   table([{t:'#'},{t:'Manager'},{t:'Captain Pts',cls:'r'}],rows);
});

addTab('green','Green Arrow King',()=>{
 const rows=D.green.map((r,i)=>({_cls:r.excluded?'excl':'',cells:[i+1,r.manager+(r.excluded?' <span class="pill">top-5 excl.</span>':''),`<b>${r.arrows}</b>`]}));
 return '<h2>Green Arrow King — ₹3,000</h2><p class="note">Gameweeks where overall (global) rank improved. Top-5 overall excluded at settlement.</p>'+
   table([{t:'#'},{t:'Manager'},{t:'Green Arrows',cls:'r'}],rows);
});

addTab('comeback','Comeback',()=>{
 if(!D.comeback.active) return '<h2>The Comeback — ₹4,000</h2><p class="note">Activates after GW19. The bottom 21 at that point compete for most points across GW20–GW38.</p>';
 const rows=D.comeback.board.map((r,i)=>({_cls:i==0?'top5':'',cells:[i+1,r.manager,`<b>${r.points}</b>`]}));
 return '<h2>The Comeback — ₹4,000</h2><p class="note">Bottom 21 after GW19 · most points GW20–38 wins.</p>'+
   table([{t:'#'},{t:'Manager'},{t:'Pts since GW20',cls:'r'}],rows);
});

addTab('diff','Differential Diamond',()=>{
 if(!D.differential.has_snapshots) return '<h2>Differential Diamond — ₹3,000</h2><p class="note">Ownership snapshots begin at GW1. The best sub-7.5%-owned starter haul will appear here.</p>';
 let h='<h2>Differential Diamond — ₹3,000</h2><p class="note">Biggest single-GW haul from a starter owned &lt;7.5% (ownership at deadline).</p>';
 if(D.differential.best){const b=D.differential.best;h+=`<div class="cardrow"><div class="card"><div class="k">Leader</div><div class="v">${b.points} pts</div><div class="s">${b.manager} · ${b.player} · GW${b.gw} · ${b.own}% owned</div></div></div>`}
 const rows=D.differential.board.map((r,i)=>({cells:[i+1,r.manager,r.player,'GW'+r.gw,`<b>${r.points}</b>`,r.own+'%']}));
 return h+table([{t:'#'},{t:'Manager'},{t:'Player'},{t:'GW',cls:'c'},{t:'Pts',cls:'r'},{t:'Own',cls:'r'}],rows);
});

addTab('pity','Pity',()=>{
 const rows=D.pity.map((r,i)=>({cells:[i+1,r.manager,'GW'+r.gw,`<b>${r.score}</b>`,r.hit?('-'+r.hit):'0',r.ruled_out==null?'n/a':r.ruled_out,r.valid?'<span class="up">✓</span>':'<span class="down">review</span>']}));
 return '<h2>Pity the Living Dead — ₹2,000</h2><p class="note">Lowest valid GW score (to GW30). "Valid" auto-checks hits ≤ 8 and ruled-out starters ≤ 2; admins confirm edge cases.</p>'+
   table([{t:'#'},{t:'Manager'},{t:'GW',cls:'c'},{t:'Score',cls:'r'},{t:'Hit',cls:'c'},{t:'Ruled-out',cls:'c'},{t:'Valid',cls:'c'}],rows);
});

// build UI
const tabsEl=document.getElementById('tabs'), panelsEl=document.getElementById('panels');
TABS.forEach((t,i)=>{
 const b=el('div','tab'+(i==0?' active':''),t.label); b.onclick=()=>{
   document.querySelectorAll('.tab').forEach(x=>x.classList.remove('active'));
   document.querySelectorAll('.panel').forEach(x=>x.classList.remove('active'));
   b.classList.add('active'); document.getElementById('p_'+t.id).classList.add('active');
 }; tabsEl.appendChild(b);
 const p=el('div','panel'+(i==0?' active':''),t.render()); p.id='p_'+t.id; panelsEl.appendChild(p);
});
const m=D.meta;
document.getElementById('hsub').innerHTML=`${m.n_members} managers · <span class="gold">₹1,24,000</span> pool · `+
 (m.current_gw?`through GW${m.current_gw}`:'season not started')+` · updated ${m.generated_utc}`+(m.demo?' · <b>DEMO DATA</b>':'');
document.getElementById('foot').innerHTML=`Auto-generated from the official FPL API · League ${m.league_id} · ${m.generated_utc}`+
 (m.demo?' · showing simulated numbers for design preview':'');
</script></body></html>"""

if __name__ == "__main__":
    main()
